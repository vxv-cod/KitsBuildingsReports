import os
import sys
from requests import get as requests_get
from requests_negotiate_sspi import HttpNegotiateAuth
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment


def GO(Id, NameFaileDoc, KO):
    url = f"http://tnnc-pir-app/test-kits-buildings-api/kits/kit-build-items/{Id}"
    response = requests_get(url, auth=HttpNegotiateAuth()).json()
    response.sort(key=lambda item: (str(item['StageNumber']), item['GenplanNumber']))
    # print(response)

    # Собираем данные таблицы
    data = []
    counterList = []
    temp = ''
    counter = 2
    xxx = [None, None, None]
    for i in response:
        # if i['KoItemId'] == int(KO):
        if i['KoItemId'] == int(KO):
            StageNumber = i['StageNumber']
            if StageNumber != temp:
                data.append([None] + [str(StageNumber) + ' этап'] + [None])
                temp = StageNumber
                counter += 1
                counterList.append(counter)
            xxx = [i['GenplanNumber'], i['Title'], None]
            data.append(xxx)
            counter += 1
    print(data)
    
    # Открываем шаблон
    # wb = load_workbook(filename = filename, keep_vba = True)
    wb = load_workbook(filename = 'Экспликация_ГП.xlsx')
    # Выбираем страницу
    ws = wb['Таблица1']

    # Загружаем данный в таблицу
    for i in data:
        ws.append(i)

    # Максимальное количество заполненных строк
    max_row = ws.max_row
    # Максимальное количество заполненных строк
    max_column = ws.max_column

    # '''Все колонки'''
    columns = tuple(ws.columns)
    
    # '''Выравниваем по центру первую колонку'''
    for cell in columns[0]:
        cell.alignment = Alignment(horizontal="center", vertical="center")        

    # # '''Выравниваем вертикали 2ой столбец'''
    cells_range = ws['B3':f'B{max_row}']
    for cel in cells_range:
        cel[0].alignment = Alignment(horizontal="left", vertical="center")

    # Выравниваем по центру название этапов
    for row in counterList:
        cel = ws[f'B{row}']
        cel.alignment = Alignment(horizontal='center')

    # Рисуем границы таблицы
    thins = Side(border_style="thin", color="000000")
    medium = Side(border_style="medium", color="000000")
    for col in ws.iter_cols(min_row=3, max_col=max_column, max_row=max_row):
        for cell in col:
            cell.border = Border(top=thins, bottom=thins, left=medium, right=medium)
            if cell == col[-1]:
                cell.border = Border(top=thins, bottom=medium, left=medium, right=medium)

    # Сохраняем файл
    wb.save(os.getcwd() + f"\\reports\{NameFaileDoc}")


if __name__ == "__main__":
    from rich import print
    # Id, NameFaileDoc, KO = '2306', 'ExpGP.xlsx', '1234'
    Id, NameFaileDoc, KO = '2306', 'ExpGP.xlsx', '1423'

    GO(Id, NameFaileDoc, KO)
    print("'Выполнено . . .'")
    sys.exit()